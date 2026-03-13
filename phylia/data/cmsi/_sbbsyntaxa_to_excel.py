
import pandas as _pd

import openpyxl as _openpyxl
from openpyxl import Workbook as _Workbook

from ._cmsi_syntaxa import CmsiSyntaxonTable
from ...tools.excel import dataframe_to_excelsheet



def sbbsyntaxa_to_excel(filepath=None):

    cste = SbbSynToExcel()
    wb = cste.workbook()

    if filepath:
        wb.save(filename=filepath) ##f"Syntaxa Staatsbosbebeer Catalogus_{today}.xlsx")
    ##cste._syn.to_csv(f"Syntaxa Staatsbosbebeer Catalogus_{today}.csv", index=True)

    return wb


class SbbSynToExcel:

    COLUMN_MEANINGS = {
        'LongScientificName':'Wetenschappelijke naam syntaxon',
        'LongCommonName':'Nederlandse naam syntaxon', 
        'Created':'Datum waarop het syntaxon is ingevoerd in CMSi', 
        'IsCurrent':'Syntaxon is actueel (Yes) of historisch (No)',
        'SynLevel':'Syntaxonomisch niveau van het syntaxon', 
        'SynClass':'Klasse waartoe het syntaxon behort', 
        'IsLowest':'Syntaxon op het laagste niveau (Yes) of hoger niveau (No)', 
        'IsCrossClass':'Syntaxon is klasseoverschrijdend (Yes)', 
        'CrossClassCodes':'Codes waaronder een klasseoverschrijdend syntaxon is opgenomen gescheiden door #',
        }

    def __init__(self):

        self._cst = CmsiSyntaxonTable()
        self._syn = self._cst.vegetationtypes(
            typology='sbbcat', 
            current_only=False, 
            include_mapcodes=False,
            verbose=False,
            )

    def __repr__(self):
        return f"{self.__class__.__name__} (n={len(self)})"

    def __len__(self):
        return len(self._syn)

    def workbook(self):
        """Return Excel workbook with syntaxon names."""

        workbook = _Workbook()
        workbook.remove(workbook.active) # remove default empty worksheet

        # write syntaxa
        wsname = "Syntaxa"
        sheet1 = workbook.create_sheet(wsname)
        
        ##sheet1_colwidth = self._sheetcolwidth(df=cste._syn)
        sheet1 = dataframe_to_excelsheet(dataframe=self._syn, sheet=sheet1, colwidth=None, tablename=wsname, freeze_cell='B2')

        # wrtie column meanings
        wsname = "Toelichting"
        sheet2 = workbook.create_sheet(wsname)

        # create dataframe with sheet contents
        sr = _pd.Series(self.COLUMN_MEANINGS, name='Toelichting')
        sr.index.name='Kolom'
        #df = sr.reset_index()

        sheet2 = dataframe_to_excelsheet(dataframe=sr, sheet=sheet2, colwidth=None, tablename=wsname, freeze_cell=None)

        workbook.close()
        return workbook


