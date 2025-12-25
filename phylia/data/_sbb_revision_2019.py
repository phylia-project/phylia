
import os as _os
import datetime as _dt
from importlib import resources as _resources
import pandas as _pd
from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font as _Font
from openpyxl.styles import Color as _Color, Fill as _Fill
from openpyxl.styles.alignment import Alignment as _Alignment
from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TableStyleInfo
import openpyxl as _openpyxl

from . import _data_sbb_intern
from .cmsi import vegetationtypes as _vegetationtypes

class SbbRevision2019:
    """Table with possible translations between Staatsbosbeheer 
    Catalogus and rVVN syntaxa including comments from experts.
    
    Notes
    -----
    Each line in the translation rules table contains a possible 
    translation between two syntaxa. Because a syntaxon in one 
    classification can potentially be translated to multiple syntaxa 
    in the other classification, a single syntaxon code can appear on 
    multiple lines. Therefore, this is not a one-on-one translation 
    table, but table with expert comments on the proposed translations 
    from one classification to the other.
        
    """

    ADDITIONAL_TRANSLATIONS_SBB_TO_RVVN = [
        ['01A1', 'r01Aa01a',],
        ['01A1', 'r01Aa01b',],
        ['01A2', 'r01Aa02a',],
        ['01A2', 'r01Aa02b',],
        ['01B1', 'r01Ab01b',],
        ['05D3a', 'r05Bc03',],
        ['05D3b', 'r05Bc03',],
        ['05D3c', 'r05Bc03',],
        ['08B-b', 'r10RG07',],
        ['11-j',  'r09RG10',],
        ['14/d',  'r32Ca03',],
        ['14D4',  'r14Bb02a',],
        ['14D4',  'r14Bb02',],
        ['19-h',  'r16RG08',],
        ['31D3',  'r32Ca03',],
        ['39A-f', 'r43Aa02',],
        ['40A2',  'r43Aa02',],
        ['42A-a', 'r45Aa05',],
        ['42A-b', 'r45Aa05',],
        ['42-b', 'r45Aa05',],
        ]

    CLASSIFICATION_COLUMNS = {
        'sbbcat':'code_sbb', 
        'vvn':'code_vvn_1998',
        'rvvn':'code_rvvn_2018',
        'veldgids2015':'code_veldgids_2015',
        }

    EXCELSHEET_COLUMNS= {
        'translation_id': 15,
        'code_sbb':13, 
        'code_rvvn_2018':17, 
        'cmsi_sbb_wetnaam':50, 
        'sbb_beginjaar':15, 
        'sbb_eindjaar':14,
        'code_vvn_1998':17, 
        'code_veldgids_2015':20,
        'kov_codes':13,
        'kov_verbonden':16,
        'vvbh':12,
        'sbb_in_rvvn':13,
        'rvvn_in_sbb':13,
        'opmerking_piet_schipper_2018':150,
        'opmerking_john_jansen_2018':150, 
        'opmerking_code_vvn_1998':27,
        'opmerking_code_veldgids_2015':30,
        'sbb_wetnaam_voor_2019':75,
        'sbb_nednaam_voor_2019':50,
        'sbb_wetnaam_kort_voor_2019':50,
        'cmsi_rvvn_wetnaam':65,
        }

    EXCELSHEET_COLMEANINGS = {
        'translation_id' : 'Uniek volgnummer van de vertaalregel',
        'code_sbb' : 'Code van het Staatsbosbeheer Catalogus syntaxon',
        'cmsi_sbb_wetnaam' : 'Wetenschappelijke naam van het syntaxon',
        'sbb_beginjaar' : 'Jaar waarin het sbb syntaxon is ingevoerd',
        'sbb_eindjaar' : 'Jaar waarin het sbb syntaxon is vervallen',
        'code_rvvn_2018' : 'Code van het rvvn syntaxon waarin het sbb syntaxon kan worden vertaald',
        'cmsi_rvvn_wetnaam' : 'Wetenschappelijke naam rvvn syntaxon',
        'code_vvn_1998' : 'Code van het rvvn syntaxon in de vvn uit 1998',
        'code_veldgids_2015' : 'Code van het vvn syntaxon in de veldgids rompgemeenschappen uit 2015',
        'kov_codes' : 'Codes waaronder een klasseoverschrijdend syntaxon is opgenomen',
        'kov_verbonden' : 'Code van de klasseoverschrijdende syntaxa waarmee een klasseoverschrijdend syntaxon is verbonden',
        'vvbh' : 'Vervangbaarheid van de gemeenschap op een schaal van 1 tot en met 5',
        'sbb_in_rvvn' : 'Syntaxon heeft volgens Piet Schipper een equivalent in de rVVN ("ja") of geen equivalent ("nee")',
        'rvvn_in_sbb' : 'Syntaxon heeft volgens Piet Schipper een equivalent in de Staatsbosbeheer Catalogus ("ja") of geen equivalent ("nee")',
        'opmerking_piet_schipper_2018' : 'Opmerking van Piet Schipper bij de vertaling',
        'opmerking_john_jansen_2018' : 'Opmerking van John Jansen bij de vertaling',
        'opmerking_code_vvn_1998' : 'Opmerking die in het veld code_vvn_2018 stond in plaats van een geldige syntaxoncode',
        'opmerking_code_veldgids_2015' : 'Opmerking die in het veld code_veldgids_2015 stond in plaats van een geldige syntaxoncode',
        'sbb_wetnaam_voor_2019' : 'Wetenschappelijke naam sbb syntaxon voor 2019 (indien gewijzigd)',
        'sbb_nednaam_voor_2019' : 'Nederlandse naam sbb syntaxon voor 2019 (indien gewijzigd)',
        'sbb_wetnaam_kort_voor_2019' : 'Korte wetenschappelijke naam sbb syntaxon voor 2019 (indien gewijzigd)',
        }
        
    EXCELSHEET_NAME = "SbbRevisie2019"

    def __init__(self): 
        
        # get raw table with translations
        srcfile = (_resources.files(_data_sbb_intern) / 'sbbcat_revisie_2019.csv')
        self._trans = _pd.read_csv(srcfile, encoding='latin-1')
        self._trans.index.name = "translation_id"

        # table sof syntaxon names
        self._sbbcat = self.syntaxa_sbb()
        self._rvvn = self.syntaxa_rvvn()

        # add column cmsi_sbb_wetnames
        sbbnames = self._sbbcat[['LongScientificName']].squeeze()
        self._trans['cmsi_sbb_wetnaam'] = self._trans[['code_sbb']].replace(sbbnames)

        # add column cmsi_rvvn_wetnames
        rvvn_names = self._rvvn[['LongScientificName']].squeeze()
        self._trans['cmsi_rvvn_wetnaam'] = self._trans[['code_rvvn_2018']].replace(rvvn_names)

        # add column "rvvn_in_sbb" for consistency
        mask = self._trans['opmerking_piet_schipper_2018']=='niet in SBB-typologie'
        self._trans.loc[~mask, 'rvvn_in_sbb'] = 'ja'
        self._trans.loc[mask, 'rvvn_in_sbb'] = 'nee'
        assert self._trans[self._trans['rvvn_in_sbb'].isnull()].empty


    def __len__(self):
        return len(self._trans)


    def __repr__(self):
        return f"{self.__class__.__name__} ({len(self)} translations)"


    def syntaxa_sbb(self):
        """Return table of syntaxa from the Staatsbosbeheer Catalogus."""

        # get full table of cmsi sbbcat syntaxa
        sbbcat = _vegetationtypes(
            typology='sbbcat', current_only=False, include_mapcodes=False,
            verbose=True)

        # drop confusing columns and rows
        sbbcat = sbbcat.drop(columns=['VegClas', 'GUID', 'Parent', 
            'Description', 'ModifiedBy', 'CreatedBy', 'Quality'])

        return sbbcat


    def syntaxa_rvvn(self):
        """Return table of syntaxaq from the rVVN."""

        # get full table of cmsi revision syntaxa
        rvvn = _vegetationtypes(
            typology='rvvn', current_only=False, include_mapcodes=False,
            verbose=True)

        # drop confusing columns and rows
        rvvn = rvvn.drop(columns=['VegClas', 'GUID', 'Parent', 
            'Description', 'ModifiedBy', 'CreatedBy', 'Quality'])

        return rvvn


    def syntaxa_with_multiple_entries(self, typology='sbbcat'):
        """Return table of duplicate syntaxon codes with number of 
        occurences.
        
        Parameters
        ----------
        typology : {'sbbcat', 'rvvn'}, default 'sbbcat'
            Name of the classification system.

        Returns
        -------
        pandas.Series
            Table of syntaxon codes that have multiplle entries with  
            the numbers of times each codes appears in a translation.
            
        """
        if typology=='sbbcat':
            # duplicate sbb codes
            sbbdup = self._trans[self._trans['code_sbb'].notna()]
            sbbdup = sbbdup[sbbdup.duplicated(subset=['code_sbb'], keep=False)]
            sbbdupfrq = sbbdup['code_sbb'].value_counts().sort_index()
            sbbdupfrq.name = 'sbbcat_duplicate_count'
            return sbbdupfrq

        elif typology=='rvvn':
            # duplicate rvvn codes
            revdup = self._trans[self._trans['code_rvvn_2018'].notna()]
            revdup = revdup[revdup.duplicated(subset=['code_rvvn_2018'], keep=False)]
            revdupfrq = revdup['code_rvvn_2018'].value_counts().sort_index()
            revdupfrq.name = 'rvvn_duplicate_count'
            return revdupfrq

        else:
            raise ValueError(f'Invalid classification system "{typology}".')


    def sbb_syntaxa_not_in_rvvn(self, verbose=False):
        """Return table of Staatsbosbeheer Syntaxca that have no equivalent 
        in the rVVN.
        
        Parameters
        ----------
        verbose : bool, default False
            Return only a selection of columns (False) or all columns (True).

        Returns
        -------
        pd.DataFrame
            Table of Staatsbosbeheer syntaxa.
            
        """
        mask = self._trans['sbb_in_rvvn']=='nee'
        no_equivalent = self._trans[mask].sort_values(
            ['sbb_eindjaar','code_sbb'])

        if not verbose:
            columns = ['code_sbb','cmsi_sbb_wetnaam',]
            return no_equivalent[columns]

        return no_equivalent


    def translations(self, from_sys='sbbcat', to_sys='rvvn'):
        """Return table of all possible translations between two 
        classification systems.
        
        Parameters
        ----------
        from_sys : 
            Classification system to translate from.
        to_sys :
            Cllassification system to translate to.

        Returns
        -------
        
        """
        if from_sys not in self.CLASSIFICATION_COLUMNS.keys():
            raise ValueError(f"Invalid 'from' classification {from_sys}. Classification must be in {self.CLASSIFICATION_COLUMNS}.")

        if to_sys not in self.CLASSIFICATION_COLUMNS.keys():
            raise ValueError(f"Invalid 'from' classification {to_sys}. Classification must be in {self.CLASSIFICATION_COLUMNS}.")

        # get syntaxoncode column names
        from_code = self.CLASSIFICATION_COLUMNS[from_sys]
        to_code = self.CLASSIFICATION_COLUMNS[to_sys]

        # create tranlationtable
        mask1 = self._trans[from_code].notna()
        mask2 = self._trans[to_code].notna()
        translations = self._trans[mask1&mask2][[from_code, to_code]].copy()

        # add missing translations that Piet has forgotten in 2019
        mask1 = from_sys in ['sbbcat','rvvn']
        mask2 = to_sys in ['sbbcat','rvvn']
        mask3 = from_sys != to_sys
        if mask1 & mask2 & mask3:
            for (sbbcode, rvvncode) in self.ADDITIONAL_TRANSLATIONS_SBB_TO_RVVN:
                idx = translations.index.values[-1] + 1
                translations.loc[idx, 'code_sbb'] = sbbcode
                translations.loc[idx, 'code_rvvn_2018'] = rvvncode

        return translations


    def sbb_syntaxa_namechanges(self):
        """Return table syntaxon names from Staatsbosbeheer Catalogus 
        before and after revision process in 2018 (returns only syntaxa 
        where name has changed)."""
        changed = self._trans.copy()
        mask1 = changed['cmsi_sbb_wetnaam']!=changed['sbb_wetnaam_voor_2019']
        mask2 = changed['cmsi_sbb_wetnaam'].notnull()
        mask3 = changed['sbb_wetnaam_voor_2019'].notnull()
        return changed[mask1&mask2&mask3].copy()


    def revisiontable(self):
        """Return table with revision data."""
        return self._trans


    def revisiontable_to_excel(self, fpath=None):
        """Save table with revision data and documentation to excel.
        
        Parameters
        ----------
        fpath : str, optional
            Valid filepath for Excelfile.

        Returns
        -------
        openpyxl Workbook
            In memory Excelfile with saved data.
            
        """
        
        # create workbook
        workbook = _Workbook()
        workbook.remove(workbook.active) # remove default empty worksheet

        # create first sheet with data
        sheet = workbook.create_sheet(self.EXCELSHEET_NAME)
        sheet = self._excelsheet_translations(sheet)

        # create second sheet with column meanings
        sheet2 = workbook.create_sheet('Toelichting')
        sheet2 = self._excelsheet_documentation(sheet2)

        if fpath:

            # add date to filepath
            today = _dt.datetime.now().strftime('%y%m%d')
            fdir, file = _os.path.split(fpath)
            fname, ext = _os.path.splitext(file)
            fpath = f"{fdir}\\{fname}_v{today}.xlsx"

            # save excelfile
            workbook.save(filename=fpath)
            workbook.close()
        
        return workbook


    def _excelsheet_translations(self, sheet):
        """Return worksheet with translation rules."""
        
        data = self._trans.reset_index()
        colwidth = self.EXCELSHEET_COLUMNS

        # compare columns in data table and excelwidth columns
        table_columns = list(data)
        excel_columns = self.EXCELSHEET_COLUMNS.keys()       
        self._excelsheet_check_columns(table_columns, excel_columns)
        
        # write column names       
        for icol, colname in enumerate(excel_columns):
            cellref = sheet.cell(row=1, column=icol+1)
            cellref.value = colname

        # write data
        for irow, rowname in enumerate(data.index.values):
            for icol, colname in enumerate(excel_columns):
                cell = sheet.cell(row=irow+2, column=icol+1)

                value = data.loc[rowname, colname]
                if not _pd.isnull(value):
                    cell.value = str(value)

                # set font size and aligment
                cell.font = _Font(size=10)
                cell.alignment = _Alignment(horizontal='left', vertical='center')

        # set column width
        for i in range(1, sheet.max_column+1):
            cell = sheet.cell(row=1, column=i)
            sheet.column_dimensions[cell.column_letter].width = colwidth[cell.value]
            cell.font = _Font(bold=True, color='FFFFFFFF')
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # freeze
        sheet.freeze_panes = sheet['D2']

        # add table
        mediumstyle = _TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
        maxcol_letter =_openpyxl.utils.get_column_letter(sheet.max_column)
        table = _Table(displayName='vertalingen', ref=f'A1:{maxcol_letter}{sheet.max_row}', 
            tableStyleInfo=mediumstyle)
        sheet.add_table(table)

        return sheet


    def _excelsheet_documentation(self, sheet):
        """Return worksheet with collumn meanings."""

        data = self._trans.reset_index()
        colwidth = self.EXCELSHEET_COLMEANINGS
        
        # compare columns in data table and excelwidth columns
        table_columns = list(data)
        excel_columns = self.EXCELSHEET_COLMEANINGS.keys()
        self._excelsheet_check_columns(table_columns, excel_columns)

        # write data
        row = 1
        cell = sheet.cell(row=row, column=1)
        cell.value = f'Toelichting'
        cell.font = _Font(bold=True)

        for textrow in [
            f'Het werkblad "{self.EXCELSHEET_NAME}" documenteert de integratie van de Staatsbosbeheer Catalogus en de revisie van de ',
            f'Vegetatie van Nederland in 2018 en 2019. Dit werk werk is uitgevoerd door Piet Schipper en Gerard Eppink (Staatsbosbeheer).',
            f'John Jansen (WENNR) heeft de voorgestelde vertalingen beoordeeld en van commentaar voorzien.',
            f'Iedere regel van het werkblad geeft een vertaling van een syntaxon uit de Staatsbosbeheer Catalogus naar een syntaxon ',
            f'uit de revisie van de vegetatie van Nederland. Omdat een syntaxon in de ene classificatie in principe vertaald kan worden',
            f'naar meerdere syntaxa in de andere classificatie, kunnen syntaxoncodes op meerdere regels voorkomen.',
            f'De gegevens in het werkblad zijn een vertaling van werk dat in 2018 is verricht. Fouten en inconsistensies zijn niet ',
            f'systematisch gecorrigeerd. Bij twijfel zijn de gegevens in de database CMSi in principe leidend.',
            ]:
            row=row+1
            cell = sheet.cell(row=row, column=1)
            cell.value = textrow

        # write collumn names
        row = row+2
        cell = sheet.cell(row=row, column=1)
        cell.value = 'Kolomnaam'
        cell.font = _Font(bold=True)
        cell = sheet.cell(row=row, column=2)
        cell.value = 'Betekenis'
        cell.font = _Font(bold=True)

        # write data
        row = row+1
        for irow, colname in enumerate(self.EXCELSHEET_COLUMNS.keys()):
            cell = sheet.cell(row=row+irow, column=1)
            cell.value = colname
            cell = sheet.cell(row=row+irow, column=2)
            cell.value = self.EXCELSHEET_COLMEANINGS[colname]
        row = row+irow+1
    
        today = _dt.datetime.now().strftime('%d-%m-%Y')
        for textrow in [
            f'Dit bestand is afgeleid uit het "werkblad SBB 2018" van het Excelbestand "revisie sbb-catalogus_PS_aug_2019_update_PS_9e.xlsx"',
            f'Neem voor een nadere toelichting op dit bestand contact op met Thomas de Meij (Staatsbosbeheer).',
            f'Dit bestand is gemaakt op {today}.'
            ]:
            row=row+1
            cell = sheet.cell(row=row, column=1)
            cell.value = textrow

        # set column width
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 200

        return sheet


    def _excelsheet_check_columns(self, table_columns, excel_columns):
        """Check for presence of necessary columns and missing columns."""

        missing_cols = [x for x in table_columns if x not in excel_columns]
        if missing_cols:
            missing_string = '"' + '", "'.join(missing_cols) + '"'
            raise ValueError((f"Missing column widths for columns {missing_string}."))
        unexpected_cols = [x for x in excel_columns if x not in table_columns]
        if unexpected_cols:
            unexpected_string = '"' + '", "'.join(unexpected_cols) + '"'
            raise ValueError((f"Column widths given for unexpected columns {unexpected_string}."))
