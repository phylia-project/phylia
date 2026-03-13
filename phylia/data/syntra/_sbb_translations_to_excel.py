

import datetime as _dt
import pandas as _pd

from phylia.data.syntra import SyntaxonTranslator
import phylia

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Color, Fill
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def sbbtranslations_to_excel(lowest_only=False, include_subass=True, fpath=None):

    translator = SbbTranslationsToExcel()
    wb = translator.workbook(lowest_only=lowest_only, 
        include_subass=include_subass)
    if fpath:
        wb.save(fpath)
    return wb


class SbbTranslationsToExcel:
    """Write formatted syntaxon translation tables to Excel."""

    COLWIDTH_SBB = {
        'Code':10,
        'LongScientificName': 75,
        'LongCommonName': 75,
        'Created':12,
        'Modified': 12,
        'IsCurrent':12,
        'SynLevel':17,
        'SynClass':12,
        'IsLowest':12,
        'IsCrossClass':14,
        'CrossClassCodes':18,
        'RevisieIdentiek':16,
        'RevisieVertaling':35,
        'RevisieVertalingHistorisch':23,
        'RevisieVertalingHistorischCount':23,
        'SbbTerugvertaling':24,
        'RevisieVertalingCount':22,
        'SbbTerugvertalingCount':24,
        'RevisieVertalingLevel' : 22, 
        'RevisieVertalingIsLowest': 24,
        'ShortScientificName':75,
        'ShortCommonName':75,
        }


    COLWIDTH_RVVN = {
        'Code':10,
        'LongScientificName':75,
        'LongCommonName':75,
        'Created':12,
        'Modified':12,
        'IsCurrent':12,
        'SynLevel':15,
        'SynClass':12,
        'IsLowest':12,
        'SbbIdentiek':16,
        'SbbVertaling':20,
        'SbbVertalingHistorisch':23,
        'SbbVertalingHistorischCount':23,
        'RevisieTerugvertaling':24,
        'SbbVertalingCount':25,
        'RevisieTerugvertalingCount':26,
        'SbbVertalingLevel':22,
        'SbbVertalingIsLowest':24,
        'ShortScientificName':75,
        'ShortCommonName':75,
        }


    COLMEANINGS_SBB = {
        'Code' : 'Code van het Staatsbosbeheer syntaxon in CMSI',
        'LongScientificName' : 'Volledige wetenschappelijke naam van het syntaxon',
        'LongCommonName' : 'Volledige nederlandse naam van het syntaxon',
        'Created' : 'Datum waarop het syntaxon is ingevoerd in CMSi (altijd 2016 of later)',
        'IsCurrent' : 'Syntaxon is actueel (Yes) of vervallen (No)',
        'SynLevel' : 'Syntaxonomisch niveau van het syntaxon',
        'SynClass' : 'Klasse waartoe het syntaxon behoort',
        'IsLowest' : 'Syntaxon is van het laagste syntaxonomische niveau',
        'IsCrossClass' : 'Syntaxon is een klasseovewrschrijdend syntaxon',
        'CrossClassCodes' : 'Codes waaaronder het klasseoverschrijdend syntaxon is opgenomen in CMSi',
        'RevisieIdentiek' : 'Staatsbosbeheer syntaxon is identiek aan een RVVN syntaxon',
        'RevisieVertaling' : 'Codes van de Revisie syntaxa waar het Staatsbosbeheer syntaxon naar toe kan worden vertaald',
        'RevisieVertalingHistorisch' : 'Codes van de historische Revisie syntaxa waar het Staatsbosbeheer syntaxon naar toe kan worden vertaald',
        'RevisieVertalingHistorischCount' : '',
        'SbbTerugvertaling' : 'Codes van de Staatsbosbeheer syntaxa waarnaar de syntaxa uit "RevisieVertaling" kunnen worden terugvertaald',
        'RevisieVertalingCount' : 'Aantal codes is het veld RevisieVertaling',
        'SbbTerugvertalingCount' : 'Aantal codes in het veld SbbTerugvertaling',
        'ShortScientificName' : 'Wetenschappelijke naam van het Staatsbosbeheer syntaxon in CMSi',
        'ShortCommonName' : 'Nederlandse naam van het Staatsbosbeheer syntaxon in CMSi',
        'Modified' : 'Datum waarop de gegevens van het syntaxon voor het laatst zijn gewijzigd in CMSi',
        'RevisieVertalingLevel' : 'Syntaxonomisch niveau van het RVVN syntaxon in kolom "RevisieVertaling"',
        'RevisieVertalingIsLowest' : 'Revisie syntaxon in "RevisieVertaling" is van het laagste syntaxonomische niveau',
        }


    COLMEANINGS_RVVN = {
        'Code' : 'Code van het Revisie syntaxon in CMSI',
        'LongScientificName' : 'Volledige wetenschappelijke naam van het syntaxon',
        'LongCommonName' : 'Volledige nederlandse naam van het syntaxon',
        'Created' : 'Datum waarop het syntaxon is ingevoerd in CMSi (altijd 2016 of later)',
        'Modified' : 'Datum waarop de gegevens van het syntaxon voor het laatst zijn gewijzigd in CMSi',
        'IsCurrent' : 'Syntaxon is actueel (Yes) of vervallen (No). No komt niet bij bij Revisie syntaxa',
        'SynLevel' : 'Syntaxonomisch niveau van het syntaxon',
        'SynClass' : 'Klasse waartoe het syntaxon behoort',
        'IsLowest' : 'Syntaxon is van het laagste syntaxonomische niveau',
        'SbbIdentiek' :  'Syntaxon is identiek aan een Staatsbowsbeheer syntaxon',
        'SbbVertaling' : 'Codes van de Staatsbosbeheer syntaxa waar het Revisie syntaxon naar toe kan worden vertaald',
        'SbbVertalingHistorisch' : 'Codes van de historische Staatsbosbeheer syntaxa waar het Revisie syntaxon naar toe kan worden vertaald',
        'SbbVertalingHistorischCount' : '',
        'RevisieTerugvertaling' : 'Codes van de Revisie syntaxa waarnaar de syntaxa uit "SbbVertaling" kunnen worden terugvertaald',
        'SbbVertalingCount' : 'Aantal codes is het veld SbbVertaling',
        'RevisieTerugvertalingCount' : 'Aantal codes in het veld RevisieTerugvertaling',
        'SbbVertalingLevel' : 'Syntaxonomisch niveau van het RVVN syntaxon in kolom "SbbVertaling"',
        'SbbVertalingIsLowest' : 'Staatsbosbeheer syntaxon in "SbbVertaling" is van het laagste syntaxonomische niveau',
        'ShortScientificName' : 'Wetenschappelijke naam van het Revisie syntaxon in CMSi',
        'ShortCommonName' : 'Nederlandse naam van het Revisie syntaxon in CMSi',
        }

    WORKSHEET_NAME_SBB = 'vertaaltabel_sbb'
    WORKSHEET_NAME_RVVN = 'vertaaltabel_rvvn'
    WORKSHEET_NAME_DOC = 'Toelichting'


    def __init__(self):

        self._translator = SyntaxonTranslator()

       
    def __repr__(self):
        return f"{self.__class__.__name__}"


    def sbb_to_rvvn(self, lowest_only=False, include_subass=True):
        """Return translations from sbb to rvvn syntaxa."""
        sbb = self._translator.translate_sbb_to_rvvn(
            lowest_only=lowest_only, include_subass=include_subass)
        sbb = sbb.sort_values('Code').reset_index(drop=False)
        missing_cols, unexpected_cols = self._check_columns(sbb, self.COLWIDTH_SBB)
        missing_cols, unexpected_cols = self._check_columns(sbb, self.COLMEANINGS_SBB)
        return sbb


    def rvvn_to_sbb(self, lowest_only=False, include_subass=True):
        """Return translations from rvvn to sbb syntaxa."""
        rvvn = self._translator.translate_rvvn_to_sbb(
            lowest_only=lowest_only, include_subass=include_subass)
        rvvn = rvvn.sort_values('Code').reset_index(drop=False)
        missing_cols, unexpected_cols = self._check_columns(rvvn, self.COLWIDTH_RVVN)
        missing_cols, unexpected_cols = self._check_columns(rvvn, self.COLMEANINGS_RVVN)
        return rvvn


    def workbook(self, lowest_only=False, include_subass=True):

        sbb = self.sbb_to_rvvn(
            lowest_only=lowest_only, 
            include_subass=include_subass,
            )

        rvvn = self.rvvn_to_sbb(
            lowest_only=lowest_only, 
            include_subass=include_subass,
            )

        # create workbook
        workbook = Workbook()
        workbook.remove(workbook.active) # remove default empty worksheet
        sheet1 = workbook.create_sheet(self.WORKSHEET_NAME_SBB)
        sheet2 = workbook.create_sheet(self.WORKSHEET_NAME_RVVN)
        sheet3 = workbook.create_sheet(self.WORKSHEET_NAME_DOC)

        sheet1 = self._write_sbb(sheet1, sbb)
        sheet2 = self._write_rvvn(sheet2, rvvn)
        sheet3 = self._write_colmeanings(sheet3)

        workbook.close()
        return workbook


    def _check_columns(self, table_columns, excel_columns):
        """Check for presence of necessary columns and missing columns."""

        missing_cols = [x for x in table_columns if x not in excel_columns]
        if missing_cols:
            missing_string = '"' + '", "'.join(missing_cols) + '"'
            raise ValueError((f"Missing columns {missing_string}."))

        unexpected_cols = [x for x in excel_columns if x not in table_columns]
        if unexpected_cols:
            unexpected_string = '"' + '", "'.join(unexpected_cols) + '"'
            raise ValueError((f"Unexpected columns {unexpected_string}."))

        return missing_cols, unexpected_cols


    def _write_sbb(self, sheet, data):

        # write column names
        for i, colname in enumerate(self.COLWIDTH_SBB.keys()):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = colname

        # write data
        for irow, idx in enumerate(data.index.values):
            for icol, colname in enumerate(self.COLWIDTH_SBB.keys()):
                cell = sheet.cell(row=irow+2, column=icol+1)

                val = data.loc[idx,colname]
                if not _pd.isnull(val):
                    cell.value = str(val)

        # set column width
        for i in range(1, sheet.max_column+1):
            cell = sheet.cell(row=1, column=i)
            sheet.column_dimensions[cell.column_letter].width = self.COLWIDTH_SBB[cell.value]
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # define table
        mediumstyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
        maxcol_letter = openpyxl.utils.get_column_letter(sheet.max_column)
        table = Table(displayName='CMSiSBB', ref=f'A1:{maxcol_letter}{sheet.max_row}', 
            tableStyleInfo=mediumstyle)
        sheet.add_table(table)

        # freeze panes
        sheet.freeze_panes = sheet['C2']

        return sheet


    def _write_rvvn(self, sheet, data):

        # write column names
        for i, colname in enumerate(self.COLWIDTH_RVVN.keys()):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = colname

        # write data
        for irow, idx in enumerate(data.index.values):
            for icol, colname in enumerate(self.COLWIDTH_RVVN.keys()):
                cell = sheet.cell(row=irow+2, column=icol+1)

                val = data.loc[idx,colname]
                if not _pd.isnull(val):
                    cell.value = str(val)

        # set column width
        for i in range(1, sheet.max_column+1):
            cell = sheet.cell(row=1, column=i)
            sheet.column_dimensions[cell.column_letter].width = self.COLWIDTH_RVVN[cell.value]
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # define table
        mediumstyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
        maxcol_letter = openpyxl.utils.get_column_letter(sheet.max_column)
        table = Table(displayName='CMSiRev', ref=f'A1:{maxcol_letter}{sheet.max_row}', 
            tableStyleInfo=mediumstyle)
        sheet.add_table(table)

        # freeze panes
        sheet.freeze_panes = sheet['C2']

        return sheet


    def _write_colmeanings(self, sheet):

        startrow = 1

        # toelichting blad2
        cell = sheet.cell(row=startrow, column=1)
        cell.value = f'Werkblad "{self.WORKSHEET_NAME_SBB}"'
        cell.font = Font(bold=True)

        # write header
        startrow = startrow+1
        cell = sheet.cell(row=startrow, column=1)
        cell.value = 'Kolomnaam'
        cell.font = Font(bold=True)
        cell = sheet.cell(row=startrow, column=2)
        cell.value = 'Betekenis'
        cell.font = Font(bold=True)

        # write data
        startrow = startrow+1
        for irow, colname in enumerate(self.COLMEANINGS_SBB.keys()):
            cell = sheet.cell(row=startrow+irow, column=1)
            cell.value = colname
            cell = sheet.cell(row=startrow+irow, column=2)
            cell.value = self.COLMEANINGS_SBB[colname]
        startrow = startrow+irow+2

        # write toelichting blad3
        cell = sheet.cell(row=startrow, column=1)
        cell.value = f'Werkblad "{self.WORKSHEET_NAME_RVVN}"'
        cell.font = Font(bold=True)

        # write header
        startrow = startrow+1
        cell = sheet.cell(row=startrow, column=1)
        cell.value = 'Kolomnaam'
        cell.font = Font(bold=True)
        cell = sheet.cell(row=startrow, column=2)
        cell.value = 'Betekenis'
        cell.font = Font(bold=True)

        # write colmeanings
        startrow = startrow+1
        for irow, colname in enumerate(self.COLMEANINGS_RVVN.keys()):
            cell = sheet.cell(row=startrow+irow, column=1)
            cell.value = colname
            cell = sheet.cell(row=startrow+irow, column=2)
            cell.value = self.COLMEANINGS_RVVN[colname]
        startrow = startrow+irow+2

        startrow = startrow+1
        cell = sheet.cell(row=startrow, column=1)
        cell.value = f'Neem voor een nadere toelichting op dit bestand contact op met Staatsbosbeheer.'
        cell = sheet.cell(row=startrow+1, column=1)
        today = _dt.datetime.now().strftime('%d-%m-%Y')
        cell.value = f'Dit spreadsheet is gemaakt op {today}.'
        startrow = startrow+1

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 200

        return sheet

