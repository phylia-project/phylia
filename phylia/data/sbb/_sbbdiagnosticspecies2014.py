

import pandas as _pd

import openpyxl as _openpyxl
from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font as _Font
from openpyxl.styles import Color as _Color
from openpyxl.styles import Fill as _Fill
from openpyxl.styles import PatternFill as _PatternFill
from openpyxl.styles.alignment import Alignment as _Alignment
from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TableStyleInfo

from ._sbbdata import _sbbcat_diagnostic_species_2014
from ..cmsi import vegetationtypes as _cmsi_vegetationtypes


def sbbcat_diagnostic_species_2014():
    """Return table of diagnstic species for Staatsbosbeheer Catalogus 
    syntaxa. 
    
    Notes
    -----
    The source for this table was a spreadsheet created by Piet Schipper 
    that contained a list of syntaxa and diagnostic species. This 
    spreadsheet was distributed among users of the Staatsbosbeheer 
    Catalogus and when users refer to the Catalogus, they usually mean
    this spreadsheet. The spreadsheet was actively maintained till about 
    2014.
        
    """
    rawtable = _sbbcat_diagnostic_species_2014()
    dia = SbbDiagnosticSpecies2014(rawtable)
    return dia.diagnostic_species()


def sbbcat_diagnostic_species_to_excel(fpath=None):
    """Save tablke of diagnostic species to excel sheet with 
    additional sheets with explanations."""
    rawtable = _sbbcat_diagnostic_species_2014()
    dia = SbbDiagnosticSpecies2014(rawtable)
    return dia.to_excel(fpath=fpath)


class SbbDiagnosticSpecies2014:
    """Table of diagnostic species for the Staatsbosbeheer Catalogus
    version 2014."""

    SYNCODES = {

        'kk'    : 'kensoort van klasse',
        'kk!'   : 'kensoort van klasse (indien gemiddeld 10-20 procent bedekkend)',
        'tk'    : 'transgredierende kensoort van klasse',

        'kv'    : 'kensoort van verbond',
        'tv'    : 'transgredierende kensoort van verbond',
        'tv!'   : 'transgredierende kensoort van verbond (indien gemiddeld 10-20 procent bedekkend)',
        'tv!!'  : 'transgredierende kensoort van verbond (indien minimaal 20 procent bedekkend)',

        'ka'    : 'kensoort van associatie',
        'ka!'   : 'kensoort van associatie (indien gemiddeld 10-20 procent bedekkend)',
        'ka!!'  : 'kensoort van associatie (indien minimaal 20 procent bedekkend)',

        'ta'    : 'transgredierende kensoort van associatie',
        'ta!'   : 'transgredierende kensoort van associatie (indien gemiddeld 10-20 procent bedekkend)',
        'ta!!'  : 'transgredierende kensoort van associatie (indien minimaal 20 procent bedekkend)',

        'd'     : 'differentierende soort',
        'd!'    : 'differentierende soort (indien gemiddeld 10-20 procent bedekkend)',
        'd!!'   : 'differentierende soort (indien minimaal 20 procent bedekkend)',

        'c'     : 'constante soort',
        'c!'    : 'constante soort (indien gemiddeld 10-20 procent bedekkend)',
        'c!!'   : 'constante soort (indien minimaal 20 procent bedekkend)',

        'x'     : 'relatieve kensoort',

        'y'     : 'soort van de karakteristieke soortsgroep',
        'y!'    : 'soort van de karakteristieke soortgroep (indien gemiddeld 10-20 procent bedekkend)',
        'y!!'   : 'soort van de karakteristieke soortgroep (indien minimaal 20 procent bedekkend)',

        'z'     : 'obligaat aanwezige soort',
        'z!'    : 'obligaat aanwezige soort (indien gemiddeld 10-20 procent bedekkend)',
        'z!!'   : 'obligaat aanwezige soort (indien minimaal 20 procent bedekkend)',

        }

    SYNTAXA_ORDER = ['klasse', 'verbond', 'associatie', 'subassociatie', 'romp', 'derivaat']

    COLUMNS = [
        'syntaxon_code', 'syntaxon_kind', 'syntaxon_scientificname',
        'taxon_scientificname', 'diagnostic_value','diagnostic_value2',
        'syndif_description', 'diagnostic_uncertain', 'formation',
        'taxon_remark', 'diagnostic_remark', 
        'note1', 'note2', 'note3', 'note4', 'note5', 
        'syntaxon_vernacularname', 'taxon_vernacularname', 
        'taxon_originalname', 'taxon_corrected', 'rownr', 
        ]


    def __init__(self, diagnostic_species):

        self._diataxa = diagnostic_species   
        self._diataxa['diagnostic_value'] = _pd.Categorical(
            self._diataxa['diagnostic_value'], 
            categories=self.SYNCODES.keys(), 
            ordered=True)
        self._diataxa['syntaxon_kind'] = _pd.Categorical(
            self._diataxa['syntaxon_kind'], 
            categories=self.SYNTAXA_ORDER, ordered=True)


    def __repr__(self):
        return f'{self.__class__.__name__} (n={str(len(self))})'


    def __len__(self):
        return len(self.diagnostic_species())


    def diagnostic_species(self):
        """Return table of diagnostic species."""

        diataxa = self._diataxa

        # check for missing columns
        missing = [col for col in self.COLUMNS if col not in diataxa.columns]
        if missing:
            raise ValueError((f"Missing column names: {missing}"))

        # check for unexpected columns
        unexpected = [col for col in diataxa.columns if col not in self.COLUMNS]
        if unexpected:
            raise ValueError((f"Unexpected column names: {unexpected}"))

        # check for missing scientific names
        missing_scientificnames = diataxa[diataxa['syntaxon_scientificname'].isnull()]
        if not missing_scientificnames.empty:
            raise ValueError((f"{len(missing_scientificnames)} missing "
                f"scientific names in list of diagnosticd species."))

        mask = ~diataxa['diagnostic_value'].isnull()
        diataxa = diataxa[mask][self.COLUMNS]
        return diataxa


    def syntaxonomic_value_frequency(self):
        """Table with frequency of occurrence of diagnostic value classes."""        
        synstatus_frq = _pd.pivot_table(data=self.diagnostic_species(), 
            values='taxon_scientificname', columns='syntaxon_kind', index='diagnostic_value', 
            aggfunc='count',  fill_value=0, margins=True, margins_name='totaal',
            observed=False,
            ).fillna(0).astype('int32')
        return synstatus_frq


    def syntaxonomic_value_meaning(self):
        """Explanation of syntaxonomc value symbols."""
        df = self.diagnostic_species()[self.diagnostic_species()['diagnostic_value'].notnull()].copy()
        diafrq1 = df['diagnostic_value'].value_counts().sort_index()
        diafrq1 = diafrq1.to_frame()
        diafrq1['description']=diafrq1.index.to_series().apply(lambda x:self.SYNCODES[x])
        return diafrq1


    def cmsi_syntaxa(self):
        """Table of syntaxa with number of diagnostic species.""" 

        syntadia = self.diagnostic_species()['syntaxon_code'].value_counts().sort_index()
        syntadia = syntadia.to_frame('TaxonCount')

        cmsisynta = _cmsi_vegetationtypes(typology='sbbcat', current_only=False, include_mapcodes=False, include_crossclass=True, verbose=False)
        cmsisynta = _pd.merge(left=cmsisynta, right=syntadia,left_index=True, right_index=True, how='outer')
        cmsisynta['TaxonCount'] = cmsisynta['TaxonCount'].fillna(0).astype('int32')
        cmsisynta.index.name='SynCode'
        return cmsisynta


    def to_excel(self, fpath=None):
        """Save table of diagnostic species to spreadsheet with additional sheets.
        
        Parameters
        ----------
        fpath : str
            Vallid filepath to create Excel workbook

        Returns
        -------
        openpyxl.Workbook

                
        """
        # tables with data to write
        table1 = self.diagnostic_species()
        table1 = table1[table1['diagnostic_value'].notnull()]

        table2 = self.syntaxonomic_value_meaning().reset_index(drop=False)
        table2 = table2.rename(columns={'count':'aantal'})

        table3 = self.syntaxonomic_value_frequency().reset_index(drop=False)

        table4 = self.cmsi_syntaxa().reset_index(drop=False)

        # create workbook
        workbook = _Workbook()
        workbook.remove(workbook.active) # remove default empty worksheet
        sheet1 = workbook.create_sheet('diagnostische_soorten')
        sheet2 = workbook.create_sheet('symbolen')
        sheet3 = workbook.create_sheet('symboolgebruik')
        sheet4 = workbook.create_sheet('cmsi_syntaxa')

        colwidth3 = {
            'diagnostic_value' : 15,
            'klasse' : 15,
            'verbond' : 15, 
            'associatie' : 15, 
            'subassociatie' : 15, 
            'romp' : 15, 
            'derivaat' : 15,
            'totaal' : 15,
            }

        colwidth4 = {
            'SynCode':10,
            'LongScientificName':70, 
            'LongCommonName':70, 
            'TaxonCount':15,
            'Created':15,
            'IsCurrent':15, 
            'SynLevel':15, 
            'SynClass':15, 
            'IsLowest':15, 
            'IsCrossClass':15,
            'CrossClassCodes':20,
            }

        # write sheets
        sheet1 = self._excel_write_first_sheet(sheet1, table1)
        sheet2 = self._excel_write_second_sheet(sheet2, table2)
        sheet3 = self._excel_write_sheet(sheet3, table3, 'SBBSYNFRQ', colwidth3)
        sheet4 = self._excel_write_sheet(sheet4, table4, 'CMSISYNTAXA', colwidth4)

        # save workbook file
        if fpath:
            workbook.save(filename=fpath)
        workbook.close()
        return workbook


    def _excel_write_first_sheet(self, sheet1, table1):
        """Write sheet1 with diagnostic species."""

        # write column names
        colwidth1 = {
            'syntaxon_code':10,
            'syntaxon_kind':15,
            'syntaxon_scientificname':65,
            'taxon_scientificname':35,
            'diagnostic_value':12,
            'syndif_description':70,
            'diagnostic_uncertain': 20,
            'diagnostic_remark' : 70,
            'taxon_remark' : 70, 
            'formation':10,
            }
        for colnr, colname in enumerate(colwidth1.keys()):
            cell = sheet1.cell(row=1, column=colnr+1)
            cell.value = colname

        # write data
        for irow, idx in enumerate(table1.index.values):
            for icol, colname in enumerate(colwidth1.keys()):
                cell = sheet1.cell(row=irow+2, column=icol+1)

                # write cell value
                val = table1.loc[idx, colname]
                if not _pd.isnull(val):
                    cell.value = str(val)

                # set font size and aligment
                cell.font = _Font(size=10)
                cell.alignment = _Alignment(horizontal='left', vertical='center')

        # set column width
        for i in range(1, sheet1.max_column+1):
            cell = sheet1.cell(row=1, column=i)
            sheet1.column_dimensions[cell.column_letter].width = colwidth1[cell.value]
            cell.font = _Font(bold=True)
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # freeze
        sheet1.freeze_panes = sheet1['A2']

        # tablestyle
        mediumstyle = _TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
        maxcol_letter = _openpyxl.utils.get_column_letter(sheet1.max_column)
        table = _Table(displayName='SBBSYNTAXA', ref=f'A1:{maxcol_letter}{sheet1.max_row}', 
            tableStyleInfo=mediumstyle)
        sheet1.add_table(table)

        # make headers white
        for colnr, colname in enumerate(colwidth1.keys()):
            cell = sheet1.cell(row=1, column=colnr+1)
            cell.font = _Font(color="FFFFFFFF", bold=True)

        return sheet1


    def _excel_write_second_sheet(self, sheet2, table2):
        """Write second sheet with meaning of diagnostic values."""

        # write column names
        colwidth2 = {
            'diagnostic_value':10, 
            'aantal':10,
            'description':70,
            }
        for colnr, colname in enumerate(colwidth2.keys()):
            cell = sheet2.cell(row=1, column=colnr+1)
            cell.value = colname
            cell.font = _Font(size=10, bold=True)
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # write data
        for irow, idx in enumerate(table2.index.values):
            for icol, colname in enumerate(colwidth2.keys()):
                cell = sheet2.cell(row=irow+2, column=icol+1)

                # write cell value
                val = table2.loc[idx, colname]
                if not _pd.isnull(val):
                    cell.value = str(val)

                # set font size and aligment
                cell.font = _Font(size=10)
                cell.alignment = _Alignment(horizontal='left', vertical='center')

        # set column width
        for i in range(1, sheet2.max_column+1):
            cell = sheet2.cell(row=1, column=i)
            sheet2.column_dimensions[cell.column_letter].width = colwidth2[cell.value]
            cell.font = _Font(bold=True)
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # freeze
        sheet2.freeze_panes = sheet2['A2']
        return sheet2


    def _excel_write_sheet(self, sheet, table, tablename, colwidth):
        """Write sheet1 with diagnostic species."""

        # write column names
        for colnr, colname in enumerate(colwidth.keys()):
            cell = sheet.cell(row=1, column=colnr+1)
            cell.value = colname
            cell.font = _Font(size=10, bold=True)
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # write data
        for irow, idx in enumerate(table.index.values):
            for icol, colname in enumerate(colwidth.keys()):
                cell = sheet.cell(row=irow+2, column=icol+1)

                # write cell value
                val = table.loc[idx, colname]
                if not _pd.isnull(val):
                    cell.value = str(val)

                # set font size and aligment
                cell.font = _Font(size=10)
                cell.alignment = _Alignment(horizontal='left', vertical='center')

        # set column width
        for i in range(1, sheet.max_column+1):
            cell = sheet.cell(row=1, column=i)
            sheet.column_dimensions[cell.column_letter].width = colwidth[cell.value]
            cell.font = _Font(bold=True)
            cell.alignment = _Alignment(horizontal='left', vertical='center')

        # freeze
        sheet.freeze_panes = sheet['A2']

        # tablestyle
        mediumstyle = _TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
        maxcol_letter = _openpyxl.utils.get_column_letter(sheet.max_column)
        table = _Table(displayName=tablename, ref=f'A1:{maxcol_letter}{sheet.max_row}', 
            tableStyleInfo=mediumstyle)
        sheet.add_table(table)

        # make headers white
        for colnr, colname in enumerate(colwidth.keys()):
            cell = sheet.cell(row=1, column=colnr+1)
            cell.font = _Font(color="FFFFFFFF", bold=True)

        return sheet

