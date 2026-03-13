
import openpyxl as _openpyxl
from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font as _Font
from openpyxl.styles import Color as _Color
from openpyxl.styles import Fill as _Fill
from openpyxl.styles.alignment import Alignment as _Alignment
#from openpyxl import styles as _styles
from openpyxl.worksheet.table import Table as _Table
from openpyxl.worksheet.table import TableStyleInfo as _TableStyleInfo
#from openpyxl import utils as _utils

import pandas as _pd


def dataframe_to_excelsheet(dataframe, sheet=None, colwidth=None, 
    colminwidth=12, colmaxwidth=55, tablename=None, freeze_cell=None):
    """Write pandas DataFrame to Excel sheet."""

    data = dataframe.reset_index(drop=False)

    if not sheet:
        workbook = _Workbook()
        ## workbook.remove(workbook.active) # remove default empty worksheet
        sheet = workbook.active ##create_sheet(wsname)

    if isinstance(sheet, str):
        sheetname = sheet
        workbook = _Workbook()
        sheet = workbook.active ##create_sheet(wsname)
        sheet.title = sheetname

    # write column names
    for i, colname in enumerate(data.columns):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = colname

    # write data
    for irow, idx in enumerate(data.index.values):
        for icol, colname in enumerate(data.columns):
            cell = sheet.cell(row=irow+2, column=icol+1)

            content = data.loc[idx, colname]
            if not _pd.isnull(content):
                cell.value = str(content)

    # format column header
    for i in range(1, sheet.max_column+1):
        cell = sheet.cell(row=1, column=i)
        cell.font = _Font(bold=True, color='FFFFFFFF')
        cell.alignment = _Alignment(horizontal='left', vertical='center')

    # set column width
    sheet = autofit_columns(sheet, minwidth=colminwidth, maxwidth=colmaxwidth)

    # define table
    mediumstyle = _TableStyleInfo(name='TableStyleMedium2', showRowStripes=False)
    maxcol_letter = _openpyxl.utils.get_column_letter(sheet.max_column)
    table = _Table(displayName=tablename, ref=f'A1:{maxcol_letter}{sheet.max_row}', 
        tableStyleInfo=mediumstyle)
    sheet.add_table(table)

    # freeze panes
    if freeze_cell:
        sheet.freeze_panes = sheet[freeze_cell]

    return sheet


def autofit_columns(ws, minwidth=12, maxwidth=55, padding=1.2):
    """
    Adjust column widths to fit content
    """
    for column_cells in ws.columns:

        # get best column width
        colwidth = max(len(str(cell.value)) for cell in column_cells) * padding
        ##colwidth = (colwidth + 0) * padding
        if colwidth<minwidth:
            colwidth = minwidth
        if colwidth>maxwidth:
            colwidth = maxwidth

        # set column width
        column_letter = _openpyxl.utils.get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = colwidth

    return ws
